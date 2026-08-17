import openpyxl
from io import BytesIO
from datetime import datetime

from django.db.models import Q, Sum, Count
from django.http import HttpResponse
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend

from accounts.models import User
from .models import (
    FdsFeeStructure, FdsBatch, FdsEnquiry, FdsTrial,
    FdsStudent, FdsWeddingGroup, FdsAttendance, FdsFeesCollection
)
from .serializers import (
    FdsFeeStructureSerializer, FdsBatchSerializer,
    FdsEnquirySerializer, FdsTrialSerializer,
    FdsStudentSerializer, FdsWeddingGroupSerializer,
    FdsAttendanceSerializer, FdsAttendanceBulkSerializer,
    FdsFeesCollectionSerializer
)


from accounts.permissions import has_dynamic_permission

# ── Permission helper ─────────────────────────────────────────────

def has_fds_permission(user, *perms):
    """Check if user has any of the given FDS permissions."""
    return any(has_dynamic_permission(user, p) for p in perms)

def fds_read_only(user):
    return has_fds_permission(user, 'fds:admin', 'fds:view')


def fds_admin(user):
    return has_fds_permission(user, 'fds:admin')


def fds_fees_access(user):
    return has_fds_permission(user, 'fds:admin', 'fds_fees:view')


# ── Fee Structure ────────────────────────────────────────────────

class FdsFeeStructureViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = FdsFeeStructureSerializer
    queryset = FdsFeeStructure.objects.all()

    def get_queryset(self):
        if not fds_read_only(self.request.user) and not fds_fees_access(self.request.user):
            return FdsFeeStructure.objects.none()
        qs = FdsFeeStructure.objects.all()
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == 'true')
        return qs

    def check_write_permission(self):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")

    def create(self, request, *args, **kwargs):
        self.check_write_permission()
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        self.check_write_permission()
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        self.check_write_permission()
        return super().destroy(request, *args, **kwargs)


# ── Batch ────────────────────────────────────────────────────────

class FdsBatchViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = FdsBatchSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['class_category', 'batch_type', 'status', 'trainer']
    search_fields = ['name']
    ordering_fields = ['name', 'class_category', 'created_at']
    ordering = ['class_category', 'name']

    def get_queryset(self):
        if not fds_read_only(self.request.user):
            return FdsBatch.objects.none()
        return FdsBatch.objects.select_related('trainer').all()

    def check_write_permission(self):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")

    def create(self, request, *args, **kwargs):
        self.check_write_permission()
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        self.check_write_permission()
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        self.check_write_permission()
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['get'])
    def students(self, request, pk=None):
        """List all students in this batch."""
        batch = self.get_object()
        students = FdsStudent.objects.filter(batch=batch, is_active=True)
        serializer = FdsStudentSerializer(students, many=True)
        return Response(serializer.data)


# ── Enquiry ──────────────────────────────────────────────────────

class FdsEnquiryViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = FdsEnquirySerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'class_interest', 'source', 'joined']
    search_fields = ['name', 'phone', 'whatsapp_no', 'enquiry_id', 'location']
    ordering_fields = ['date', 'name', 'created_at', 'status']
    ordering = ['-date']

    def get_queryset(self):
        if not fds_read_only(self.request.user):
            return FdsEnquiry.objects.none()
        qs = FdsEnquiry.objects.select_related('created_by').prefetch_related('trials')

        # Date range filter
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)

        # Follow-up due filter
        follow_up_due = self.request.query_params.get('follow_up_due')
        if follow_up_due:
            from django.utils import timezone
            today = timezone.now().date()
            qs = qs.filter(
                Q(follow_up_1__lte=today) | Q(follow_up_2__lte=today),
                ~Q(status__in=['CONVERTED', 'LOST'])
            )

        return qs

    def perform_create(self, serializer):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        serializer.save(created_by=self.request.user)

    def update(self, request, *args, **kwargs):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        qs = self.get_queryset()
        from django.utils import timezone
        today = timezone.now().date()
        return Response({
            'total': qs.count(),
            'new': qs.filter(status='NEW').count(),
            'contacted': qs.filter(status='CONTACTED').count(),
            'trial_scheduled': qs.filter(status='TRIAL_SCHEDULED').count(),
            'converted': qs.filter(status='CONVERTED').count(),
            'lost': qs.filter(status='LOST').count(),
            'by_class': {
                'dance': qs.filter(class_interest='DANCE').count(),
                'zumba': qs.filter(class_interest='ZUMBA').count(),
                'yoga': qs.filter(class_interest='YOGA').count(),
            },
            'follow_up_due': qs.filter(
                Q(follow_up_1__lte=today) | Q(follow_up_2__lte=today),
                ~Q(status__in=['CONVERTED', 'LOST'])
            ).count(),
        })

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export enquiries to Excel matching the original template."""
        if not fds_read_only(request.user):
            return Response(status=403)
        qs = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ENQUIRY"
        headers = [
            'Enquiry ID', 'Date', 'Name', 'Location', 'Age', 'Source',
            'Phone', "What's App no.", 'Preffered Timing', 'Status',
            'Follow Up1', 'Follow Up 2', 'Joined Or Not', 'Remarks / Concerns',
            'Class Interest'
        ]
        ws.append(headers)
        for e in qs:
            ws.append([
                e.enquiry_id, e.date.strftime('%d/%m/%Y') if e.date else '',
                e.name, e.location or '', e.age or '',
                e.get_source_display(), e.phone or '', e.whatsapp_no or '',
                e.preferred_timing or '', e.get_status_display(),
                e.follow_up_1.strftime('%d/%m/%Y') if e.follow_up_1 else '',
                e.follow_up_2.strftime('%d/%m/%Y') if e.follow_up_2 else '',
                'Yes' if e.joined else 'No',
                e.remarks or '', e.get_class_interest_display(),
            ])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="FDS_Enquiries.xlsx"'
        return resp

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Import enquiries from Excel."""
        if not fds_admin(request.user):
            return Response({"error": "FDS admin permission required."}, status=403)
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file provided."}, status=400)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created, skipped = 0, 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # Skip header
                if not row[2]:  # Name required
                    skipped += 1
                    continue
                raw_date = row[1]
                if isinstance(raw_date, datetime):
                    date = raw_date.date()
                elif isinstance(raw_date, str):
                    try:
                        date = datetime.strptime(raw_date, '%d/%m/%Y').date()
                    except Exception:
                        date = datetime.now().date()
                else:
                    date = datetime.now().date()
                FdsEnquiry.objects.create(
                    date=date,
                    name=str(row[2]).strip(),
                    location=str(row[3]).strip() if row[3] else '',
                    age=int(row[4]) if row[4] and str(row[4]).isdigit() else None,
                    phone=str(row[6]).strip() if row[6] else '',
                    whatsapp_no=str(row[7]).strip() if row[7] else '',
                    preferred_timing=str(row[8]).strip() if row[8] else '',
                    remarks=str(row[13]).strip() if row[13] else '',
                    created_by=request.user,
                )
                created += 1
            return Response({"created": created, "skipped": skipped})
        except Exception as e:
            return Response({"error": str(e)}, status=400)


# ── Trial ────────────────────────────────────────────────────────

class FdsTrialViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = FdsTrialSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'class_category', 'converted']
    search_fields = ['name', 'phone', 'trial_id']
    ordering_fields = ['date', 'name', 'created_at']
    ordering = ['-date']

    def get_queryset(self):
        if not fds_read_only(self.request.user):
            return FdsTrial.objects.none()
        qs = FdsTrial.objects.select_related('enquiry', 'conducted_by', 'created_by')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        follow_up_due = self.request.query_params.get('follow_up_due')
        if follow_up_due:
            from django.utils import timezone
            today = timezone.now().date()
            qs = qs.filter(follow_up_date__lte=today, converted=False)
        return qs

    def perform_create(self, serializer):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        serializer.save(created_by=self.request.user)

    def update(self, request, *args, **kwargs):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        qs = self.get_queryset()
        return Response({
            'total': qs.count(),
            'scheduled': qs.filter(status='SCHEDULED').count(),
            'completed': qs.filter(status='COMPLETED').count(),
            'no_show': qs.filter(status='NO_SHOW').count(),
            'cancelled': qs.filter(status='CANCELLED').count(),
            'converted': qs.filter(converted=True).count(),
            'conversion_rate': round(
                qs.filter(converted=True).count() / qs.filter(status='COMPLETED').count() * 100, 1
            ) if qs.filter(status='COMPLETED').count() > 0 else 0,
            'avg_rating': qs.filter(
                trainer_rating__isnull=False
            ).aggregate(avg=Sum('trainer_rating') / Count('id'))['avg'] or 0,
            'by_class': {
                'dance': qs.filter(class_category='DANCE').count(),
                'zumba': qs.filter(class_category='ZUMBA').count(),
                'yoga': qs.filter(class_category='YOGA').count(),
            },
        })

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        if not fds_read_only(request.user):
            return Response(status=403)
        qs = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TRIAL"
        ws.append([
            'TRIAL ID', 'DATE', 'TIME', 'NAME', 'AGE', 'PHONE', 'LOCATION',
            'CLASS CATEGORY', 'FEE QUOTED', 'FEEDBACK', 'TRAINER RATING(IN 5)',
            'STATUS', 'CONVERTED', 'JOIN DATE', 'FOLLOW UP DATE', 'REMARKS'
        ])
        for t in qs:
            ws.append([
                t.trial_id,
                t.date.strftime('%d/%m/%Y') if t.date else '',
                t.time.strftime('%H:%M') if t.time else '',
                t.name, t.age or '', t.phone or '', t.location or '',
                t.get_class_category_display(),
                float(t.fee_quoted),
                t.feedback or '', t.trainer_rating or '',
                t.get_status_display(),
                'Yes' if t.converted else 'No',
                t.join_date.strftime('%d/%m/%Y') if t.join_date else '',
                t.follow_up_date.strftime('%d/%m/%Y') if t.follow_up_date else '',
                t.remarks or '',
            ])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="FDS_Trials.xlsx"'
        return resp


# ── Student ──────────────────────────────────────────────────────

class FdsStudentViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = FdsStudentSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_active', 'batch', 'batch__class_category', 'student_type', 'gender', 'media_consent', 'can_leave_alone']
    search_fields = ['name', 'student_id', 'contact_no', 'whatsapp_no', 'parent_name']
    ordering_fields = ['name', 'joining_date', 'created_at']
    ordering = ['name']

    def get_queryset(self):
        if not fds_read_only(self.request.user):
            return FdsStudent.objects.none()
        qs = FdsStudent.objects.select_related(
            'batch', 'fee_structure', 'created_by', 'enquiry', 'trial'
        ).prefetch_related('fds_attendances')

        # Class category filter (via batch)
        class_cat = self.request.query_params.get('class_category')
        if class_cat:
            qs = qs.filter(batch__class_category=class_cat)

        date_from = self.request.query_params.get('joining_date_from')
        date_to = self.request.query_params.get('joining_date_to')
        if date_from:
            qs = qs.filter(joining_date__gte=date_from)
        if date_to:
            qs = qs.filter(joining_date__lte=date_to)

        return qs

    def perform_create(self, serializer):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        serializer.save(created_by=self.request.user)

    def update(self, request, *args, **kwargs):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        if not fds_read_only(request.user):
            return Response(status=403)
        qs = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "REGISTRATION DETAILS"
        ws.append([
            'Student ID', 'Name ', 'Joining Date', 'Age &Gender', 'Parent Name ',
            'Contact No.', 'Emergency contact NO.', 'Batch/Time', 'Medical Condition',
            'Media Consent', 'Pickup Person 1 NO.', 'Can Leave alone',
            'Admission Fee Paid Date', 'Class Category', 'Fee Type'
        ])
        for s in qs:
            age_gender = f"{s.age or ''} / {s.get_gender_display() if s.gender else ''}".strip(' /')
            ws.append([
                s.student_id, s.name,
                s.joining_date.strftime('%d/%m/%Y') if s.joining_date else '',
                age_gender, s.parent_name or '',
                s.contact_no or '', s.emergency_contact_no or '',
                str(s.batch) if s.batch else '',
                s.medical_condition or '',
                'Yes' if s.media_consent else 'No',
                s.pickup_person_1_no or '',
                'Yes' if s.can_leave_alone else 'No',
                s.admission_fee_paid_date.strftime('%d/%m/%Y') if s.admission_fee_paid_date else '',
                s.class_category or '',
                str(s.fee_structure) if s.fee_structure else '',
            ])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="FDS_Students.xlsx"'
        return resp

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        if not fds_admin(request.user):
            return Response({"error": "FDS admin permission required."}, status=403)
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file provided."}, status=400)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created, skipped = 0, 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if not row[1]:
                    skipped += 1
                    continue
                raw_date = row[2]
                if isinstance(raw_date, datetime):
                    joining_date = raw_date.date()
                elif isinstance(raw_date, str):
                    try:
                        joining_date = datetime.strptime(raw_date, '%d/%m/%Y').date()
                    except Exception:
                        joining_date = datetime.now().date()
                else:
                    joining_date = datetime.now().date()
                FdsStudent.objects.create(
                    name=str(row[1]).strip(),
                    joining_date=joining_date,
                    parent_name=str(row[4]).strip() if row[4] else '',
                    contact_no=str(row[5]).strip() if row[5] else '',
                    emergency_contact_no=str(row[6]).strip() if row[6] else '',
                    medical_condition=str(row[8]).strip() if row[8] else '',
                    media_consent=str(row[9]).strip().lower() in ['yes', 'true', '1'] if row[9] else False,
                    pickup_person_1_no=str(row[10]).strip() if row[10] else '',
                    can_leave_alone=str(row[11]).strip().lower() in ['yes', 'true', '1'] if row[11] else False,
                    created_by=request.user,
                )
                created += 1
            return Response({"created": created, "skipped": skipped})
        except Exception as e:
            return Response({"error": str(e)}, status=400)


# ── Wedding Group ─────────────────────────────────────────────────

class FdsWeddingGroupViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = FdsWeddingGroupSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'package_type', 'trainer']
    search_fields = ['event_name', 'lead_contact_name', 'group_id', 'lead_contact_phone']
    ordering_fields = ['event_date', 'created_at', 'event_name']
    ordering = ['-created_at']

    def get_queryset(self):
        if not fds_read_only(self.request.user):
            return FdsWeddingGroup.objects.none()
        return FdsWeddingGroup.objects.select_related('batch', 'trainer', 'created_by')

    def perform_create(self, serializer):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        serializer.save(created_by=self.request.user)

    def update(self, request, *args, **kwargs):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        return super().destroy(request, *args, **kwargs)


# ── Attendance ────────────────────────────────────────────────────

class FdsAttendanceViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = FdsAttendanceSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['batch', 'date', 'status', 'class_category', 'late_arrival']
    search_fields = ['student__name', 'student__student_id']
    ordering_fields = ['date', 'class_category']
    ordering = ['-date']

    def get_queryset(self):
        if not fds_read_only(self.request.user):
            return FdsAttendance.objects.none()
        qs = FdsAttendance.objects.select_related('student', 'batch', 'marked_by')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        student_id = self.request.query_params.get('student')
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if student_id:
            qs = qs.filter(student_id=student_id)
        return qs

    def perform_create(self, serializer):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        # Auto-fill time fields from batch
        batch = serializer.validated_data.get('batch')
        if batch:
            serializer.save(
                marked_by=self.request.user,
                class_start_time=batch.time_slot_start,
                class_end_time=batch.time_slot_end,
                class_category=batch.class_category,
            )
        else:
            serializer.save(marked_by=self.request.user)

    def update(self, request, *args, **kwargs):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['post'])
    def bulk_mark(self, request):
        """Mark attendance for all students in a batch on a date."""
        if not fds_admin(request.user):
            return Response({"error": "FDS admin permission required."}, status=403)
        serializer = FdsAttendanceBulkSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        try:
            batch = FdsBatch.objects.get(id=data['batch_id'])
        except FdsBatch.DoesNotExist:
            return Response({"error": "Batch not found."}, status=404)

        created, updated = 0, 0
        for record in data['records']:
            student_id = record.get('student_id')
            att_status = record.get('status', 'PRESENT')
            late_arrival = record.get('late_arrival', False)
            notes = record.get('notes', '')
            try:
                student = FdsStudent.objects.get(id=student_id)
            except FdsStudent.DoesNotExist:
                continue
            obj, is_new = FdsAttendance.objects.update_or_create(
                student=student,
                batch=batch,
                date=data['date'],
                defaults={
                    'status': att_status,
                    'late_arrival': late_arrival,
                    'notes': notes,
                    'marked_by': request.user,
                    'class_start_time': batch.time_slot_start,
                    'class_end_time': batch.time_slot_end,
                    'class_category': batch.class_category,
                }
            )
            if is_new:
                created += 1
            else:
                updated += 1
        return Response({"created": created, "updated": updated})

    @action(detail=False, methods=['get'])
    def monthly_report(self, request):
        """Per-student monthly attendance summary for a batch."""
        if not fds_read_only(request.user):
            return Response(status=403)
        batch_id = request.query_params.get('batch_id')
        month = request.query_params.get('month')
        year = request.query_params.get('year')
        if not all([batch_id, month, year]):
            return Response({"error": "batch_id, month, and year are required."}, status=400)
        atts = FdsAttendance.objects.filter(
            batch_id=batch_id, date__month=month, date__year=year
        ).select_related('student')
        # Group by student
        report = {}
        for att in atts:
            sid = att.student.student_id
            if sid not in report:
                report[sid] = {
                    'student_id': att.student.student_id,
                    'name': att.student.name,
                    'days': {}
                }
            report[sid]['days'][str(att.date)] = att.status
        # Add summary counts
        for sid in report:
            days = report[sid]['days']
            report[sid]['present'] = sum(1 for v in days.values() if v == 'PRESENT')
            report[sid]['absent'] = sum(1 for v in days.values() if v == 'ABSENT')
            report[sid]['leave'] = sum(1 for v in days.values() if v == 'LEAVE')
            report[sid]['total'] = len(days)
            report[sid]['pct'] = round(report[sid]['present'] / report[sid]['total'] * 100, 1) if report[sid]['total'] else 0
        return Response(list(report.values()))


# ── Fees Collection ───────────────────────────────────────────────

class FdsFeesCollectionViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = FdsFeesCollectionSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'mode_of_pay', 'fees_type', 'fee_month', 'fee_year']
    search_fields = ['payment_id', 'student__name', 'student__student_id', 'wedding_group__event_name']
    ordering_fields = ['pay_date', 'paid_amount', 'created_at']
    ordering = ['-pay_date']

    def get_queryset(self):
        if not fds_fees_access(self.request.user) and not fds_read_only(self.request.user):
            return FdsFeesCollection.objects.none()
        qs = FdsFeesCollection.objects.select_related(
            'student', 'wedding_group', 'fees_type', 'collected_by'
        )
        student_id = self.request.query_params.get('student_id')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        class_cat = self.request.query_params.get('class_category')
        if student_id:
            qs = qs.filter(student_id=student_id)
        if date_from:
            qs = qs.filter(pay_date__gte=date_from)
        if date_to:
            qs = qs.filter(pay_date__lte=date_to)
        if class_cat:
            qs = qs.filter(student__batch__class_category=class_cat)
        return qs

    def perform_create(self, serializer):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        serializer.save(collected_by=self.request.user)

    def update(self, request, *args, **kwargs):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not fds_admin(self.request.user):
            self.permission_denied(self.request, message="FDS admin permission required.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Financial summary stats."""
        if not fds_fees_access(request.user) and not fds_read_only(request.user):
            return Response(status=403)
        qs = self.get_queryset()
        totals = qs.aggregate(
            total_collected=Sum('paid_amount'),
            total_billed=Sum('total_fees'),
            total_balance=Sum('balance'),
        )
        return Response({
            'total_collected': totals['total_collected'] or 0,
            'total_billed': totals['total_billed'] or 0,
            'total_balance': totals['total_balance'] or 0,
            'by_mode': {
                m[0]: qs.filter(mode_of_pay=m[0]).aggregate(t=Sum('paid_amount'))['t'] or 0
                for m in FdsFeesCollection._meta.get_field('mode_of_pay').choices
            },
            'by_status': {
                s[0]: qs.filter(status=s[0]).count()
                for s in FdsFeesCollection._meta.get_field('status').choices
            },
        })

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        if not fds_fees_access(request.user) and not fds_read_only(request.user):
            return Response(status=403)
        qs = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FEES COLLECTION"
        ws.append([
            'Student ID', 'Pay Date', 'Student name', "What's app No.",
            'Joined Date', 'Batch/Time', 'Fees Type', 'Month', 'Paid Amount',
            'Total Fees', 'Balance', 'Mode Of Pay', 'PDF LINK', 'Status/Remarks'
        ])
        for p in qs:
            student_id = p.student.student_id if p.student else (p.wedding_group.group_id if p.wedding_group else '')
            student_name = p.student.name if p.student else (p.wedding_group.event_name if p.wedding_group else '')
            whatsapp = p.student.whatsapp_no if p.student else (p.wedding_group.lead_contact_phone if p.wedding_group else '')
            joined_date = p.student.joining_date.strftime('%d/%m/%Y') if p.student and p.student.joining_date else ''
            batch_str = str(p.student.batch) if p.student and p.student.batch else ''
            import calendar
            month_str = f"{calendar.month_name[p.fee_month]} {p.fee_year}" if p.fee_month and p.fee_year else ''
            ws.append([
                student_id, p.pay_date.strftime('%d/%m/%Y') if p.pay_date else '',
                student_name, whatsapp, joined_date, batch_str,
                p.fees_type.get_category_display() if p.fees_type else '',
                month_str, float(p.paid_amount), float(p.total_fees),
                float(p.balance), p.get_mode_of_pay_display(),
                p.pdf_link or '', p.remarks or '',
            ])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="FDS_Fees_Collection.xlsx"'
        return resp


# ── Dashboard Stats ───────────────────────────────────────────────

class FdsDashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not fds_read_only(request.user) and not fds_fees_access(request.user):
            return Response({"error": "Permission denied."}, status=403)

        from django.utils import timezone
        today = timezone.now().date()
        this_month_start = today.replace(day=1)

        students = FdsStudent.objects.filter(is_active=True)
        batches = FdsBatch.objects.filter(status='ACTIVE')
        enquiries = FdsEnquiry.objects.all()
        trials = FdsTrial.objects.all()
        payments = FdsFeesCollection.objects.all()

        response = {
            'students': {
                'total_active': students.count(),
                'by_category': {
                    'dance': students.filter(batch__class_category='DANCE').count(),
                    'zumba': students.filter(batch__class_category='ZUMBA').count(),
                    'yoga': students.filter(batch__class_category='YOGA').count(),
                },
                'new_this_month': students.filter(joining_date__gte=this_month_start).count(),
            },
            'batches': {
                'total_active': batches.count(),
                'by_category': {
                    'dance': batches.filter(class_category='DANCE').count(),
                    'zumba': batches.filter(class_category='ZUMBA').count(),
                    'yoga': batches.filter(class_category='YOGA').count(),
                },
            },
            'enquiries': {
                'total': enquiries.count(),
                'new_this_week': enquiries.filter(date__gte=today - timezone.timedelta(days=7)).count(),
                'pending': enquiries.filter(status__in=['NEW', 'CONTACTED', 'TRIAL_SCHEDULED']).count(),
            },
            'trials': {
                'total': trials.count(),
                'this_week': trials.filter(date__gte=today - timezone.timedelta(days=7)).count(),
                'conversion_rate': round(
                    trials.filter(converted=True).count() / trials.filter(status='COMPLETED').count() * 100, 1
                ) if trials.filter(status='COMPLETED').count() > 0 else 0,
            },
            'fees': {
                'this_month_collected': payments.filter(
                    pay_date__gte=this_month_start
                ).aggregate(t=Sum('paid_amount'))['t'] or 0,
                'total_outstanding': payments.aggregate(t=Sum('balance'))['t'] or 0,
                'pending_count': payments.filter(status='PENDING').count(),
            },
            'wedding_groups': {
                'total_active': FdsWeddingGroup.objects.filter(status__in=['CONFIRMED', 'IN_PROGRESS']).count(),
            }
        }
        return Response(response)


# ── FDS Trainers (staff filtered by company=FDS) ──────────────────

class FdsTrainerListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not fds_read_only(request.user):
            return Response(status=403)
        trainers = User.objects.filter(company='FDS', is_active=True).values(
            'id', 'first_name', 'last_name', 'username'
        )
        data = [
            {
                'id': t['id'],
                'name': f"{t['first_name']} {t['last_name']}".strip() or t['username']
            }
            for t in trainers
        ]
        return Response(data)
