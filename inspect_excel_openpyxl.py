import openpyxl

file_path = r'b:\lp alternative\updated assets list kochi.xlsx'

print("Loading workbook...")
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"\n--- Sheet: {sheetname} ---")
    
    # Read the first 20 rows
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i >= 20:
            break
        # Print only rows that have at least one non-None value
        if any(cell is not None for cell in row):
            print(f"Row {i+1}: {row}")
