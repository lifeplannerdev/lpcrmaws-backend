from decimal import Decimal

from openpyxl import load_workbook
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from trainers.models import Student

from fees.models import FeePlanTemplate, StudentFeeAccount, FeeInstallment


class Command(BaseCommand):
    help = 'Import spreadsheet-style fee ledgers into structured fee accounts.'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str)
        parser.add_argument('--sheet', type=str, default=None)
        parser.add_argument('--company', type=str, default='FLAG')
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        xlsx_path = options['xlsx_path']
        sheet_name = options['sheet']
        company = options['company']
        dry_run = options['dry_run']

        try:
            wb = load_workbook(xlsx_path, data_only=False)
        except Exception as exc:
            raise CommandError(f'Unable to open workbook: {exc}')

        worksheets = [wb[sheet_name]] if sheet_name else wb.worksheets
        created_accounts = 0
        updated_accounts = 0
        unmatched = []

        for ws in worksheets:
            if ws.max_row < 3:
                continue

            self.stdout.write(f'Processing sheet: {ws.title}')
            template = self._get_or_create_default_template(company, ws.title, dry_run=dry_run)

            for row in ws.iter_rows(min_row=3, values_only=True):
                if not row or not row[1]:
                    continue

                student_name = str(row[1]).strip()
                course = str(row[2]).strip() if row[2] else template.course_label or ''
                student = Student.objects.filter(name__iexact=student_name, company=company).first()
                if not student:
                    unmatched.append({'sheet': ws.title, 'student_name': student_name, 'course': course})
                    continue

                entries = self._parse_payment_pairs(row)

                if dry_run:
                    created_accounts += 1
                    continue

                with transaction.atomic():
                    account, created = StudentFeeAccount.objects.get_or_create(
                        student=student,
                        defaults={
                            'company': company,
                            'template': template,
                            'plan_code': template.code,
                            'plan_name': template.name,
                            'plan_type': template.plan_type,
                            'total_due': self._compute_total_due(entries, template),
                            'registration_amount': template.registration_amount,
                            'due_day': template.due_day,
                            'source_label': ws.title,
                            'plan_snapshot': {'source_sheet': ws.title, 'source_row': self._row_as_list(row)},
                        },
                    )

                    if not created:
                        account.template = template
                        account.plan_code = template.code
                        account.plan_name = template.name
                        account.plan_type = template.plan_type
                        account.total_due = self._compute_total_due(entries, template)
                        account.registration_amount = template.registration_amount
                        account.due_day = template.due_day
                        account.source_label = ws.title
                        account.plan_snapshot = {'source_sheet': ws.title, 'source_row': self._row_as_list(row)}
                        account.save()
                        account.installments.all().delete()
                        updated_accounts += 1
                    else:
                        created_accounts += 1

                    for idx, entry in enumerate(entries, start=1):
                        FeeInstallment.objects.create(
                            account=account,
                            sequence_number=idx,
                            label=entry['label'],
                            due_date=entry['due_date'],
                            scheduled_amount=entry['amount'],
                            paid_amount=entry['paid_amount'],
                            balance_amount=max(Decimal('0'), entry['amount'] - entry['paid_amount']),
                            status=entry['status'],
                            notes=entry.get('notes', ''),
                        )
                    account.recalculate(save=True)

        self.stdout.write(self.style.SUCCESS(f'Created/updated accounts: {created_accounts + updated_accounts}'))
        if unmatched:
            self.stdout.write(self.style.WARNING(f'Unmatched rows: {len(unmatched)}'))
            for row in unmatched[:20]:
                self.stdout.write(str(row))
        else:
            self.stdout.write(self.style.SUCCESS('No unmatched rows.'))

    def _get_or_create_default_template(self, company, sheet_name, dry_run=False):
        defaults = {
            'company': company,
            'code': f'{company}-{sheet_name[:40].upper().replace(" ", "-")}-LEGACY',
            'name': sheet_name.strip(),
            'course_label': sheet_name.strip(),
            'plan_type': 'CUSTOM',
            'total_amount': Decimal('0'),
            'registration_amount': Decimal('0'),
            'due_day': 10,
            'notes': 'Legacy import template created from spreadsheet ledger.',
        }
        if dry_run:
            return FeePlanTemplate(**defaults)
        template, _ = FeePlanTemplate.objects.get_or_create(code=defaults['code'], defaults=defaults)
        return template

    def _parse_payment_pairs(self, row):
        # Spreadsheet columns are: NO, NAME, COURSE, DATE, AMOUNT, DATE, AMOUNT, ...
        entries = []
        pairs = [
            (row[3] if len(row) > 3 else None, row[4] if len(row) > 4 else None),
            (row[5] if len(row) > 5 else None, row[6] if len(row) > 6 else None),
            (row[7] if len(row) > 7 else None, row[8] if len(row) > 8 else None),
            (row[9] if len(row) > 9 else None, row[10] if len(row) > 10 else None),
        ]
        for idx, (date_value, amount_value) in enumerate(pairs, start=1):
            if not date_value or not amount_value or amount_value == '-':
                continue
            amount = self._to_decimal(amount_value)
            entries.append({
                'label': f'Ledger entry {idx}',
                'due_date': self._to_date(date_value),
                'amount': amount,
                'paid_amount': amount,
                'status': 'PAID',
                'notes': '',
            })
        return entries

    def _compute_total_due(self, entries, template):
        total = sum((entry['amount'] for entry in entries), Decimal('0'))
        if total > 0:
            return total
        return template.total_amount

    def _to_decimal(self, value):
        if isinstance(value, Decimal):
            return value
        text = str(value).replace(',', '').strip()
        return Decimal(text)

    def _to_date(self, value):
        return value.date() if hasattr(value, 'date') else value

    def _row_as_list(self, row):
        return [str(item) if item is not None else None for item in row]
