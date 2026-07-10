import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.db.models import Sum, Q, Prefetch
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import BasePermission
from core.permissions import require_permission

from fees.models import StudentFeeAccount, FeePayment, FeeAdjustment
from trainers.models import Student


class CanViewFeesGrid(BasePermission):
    def has_permission(self, request, view):
        return request.user.has_perm_name('fees_grid:view') or request.user.is_superuser


class CanExportFees(BasePermission):
    def has_permission(self, request, view):
        return request.user.has_perm_name('fees:export') or request.user.is_superuser


def get_grid_data(request):
    """Helper to fetch and shape the data based on request filters"""
    company = request.headers.get('X-Company', 'LP')
    
    # Optional filters
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    batch_name = request.query_params.get('batch')
    branch_id = request.query_params.get('branch_id')
    status = request.query_params.get('status')
    
    accounts = StudentFeeAccount.objects.filter(company=company).select_related(
        'student',
        'student__trainer__user',
        'student__branch'
    ).prefetch_related(
        Prefetch('payments', queryset=FeePayment.objects.order_by('payment_date')),
        Prefetch('adjustments', queryset=FeeAdjustment.objects.filter(adjustment_type='DISCOUNT'))
    )

    if start_date:
        accounts = accounts.filter(student__admission_date__gte=start_date)
    if end_date:
        accounts = accounts.filter(student__admission_date__lte=end_date)
    if batch_name:
        accounts = accounts.filter(student__batch=batch_name)
    if branch_id:
        accounts = accounts.filter(student__branch_id=branch_id)
    if status:
        accounts = accounts.filter(status=status)

    accounts = accounts.order_by('-student__admission_date', 'student__name')

    data = []
    max_payments = 0

    for idx, acc in enumerate(accounts, 1):
        student = acc.student
        trainer_name = student.trainer.user.get_full_name() if student.trainer else ""
        
        # Calculate discount
        discount = sum(adj.amount_delta for adj in acc.adjustments.all())

        # Get payments
        payments_list = []
        for p in acc.payments.all():
            payments_list.append({
                "amount": float(p.amount),
                "method": p.get_payment_method_display(),
                "date": p.payment_date.strftime('%Y-%m-%d') if p.payment_date else ""
            })
        
        max_payments = max(max_payments, len(payments_list))

        data.append({
            "id": acc.id,
            "sl_no": idx,
            "handled_by": trainer_name,
            "date_of_joining": student.admission_date.strftime('%Y-%m-%d') if student.admission_date else "",
            "name": student.name,
            "ph_no": student.phone_number or "",
            "parent_name": student.parent_name or "",
            "parent_no": student.parent_phone or "",
            "mail_id": student.email or "",
            "qualification": student.qualification or "",
            "campus": student.branch.name if student.branch else "",
            "mode_of_study": student.get_mode_of_study_display(),
            "preferred_country": student.preferred_country or "",
            "preferred_level": student.get_preferred_level_display() or "",
            "package_chosen": acc.plan_name or acc.plan_type,
            "total_fee": float(acc.total_due),
            "special_discount": float(discount),
            "pending": float(acc.balance_due),
            "registration_fees": float(acc.registration_amount),
            "status_of_fee": acc.get_status_display(),
            "payments": payments_list,
        })

    return data, max_payments


class FeeGridListAPIView(APIView):
    # permission_classes = [CanViewFeesGrid] # You can uncomment this once roles are assigned, keeping open for testing if needed
    
    def get(self, request, *args, **kwargs):
        data, max_payments = get_grid_data(request)
        return Response({
            "max_payments": max_payments,
            "data": data
        })


class FeeGridExportAPIView(APIView):
    # permission_classes = [CanExportFees]

    def get(self, request, *args, **kwargs):
        data, max_payments = get_grid_data(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fee Grid Export"

        # Define Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alignment_center = Alignment(horizontal='center', vertical='center')

        # Static Headers
        static_headers = [
            "SL NO", "HANDLED BY", "DATE OF JOINING", "NAME", "PH NO", "PARENT NAME", "PARENT NO",
            "MAIL ID", "QUALIFICATION", "CAMPUS", "MODE OF STUDY", "PREFERRED COUNTRY",
            "PREFERRED LEVEL", "PACKAGE CHOSEN", "TOTAL FEE", "SPECIAL DISCOUNT", "PENDING",
            "REGISTRATION FEES"
        ]

        # Write Static Headers (Row 1 and 2 merged)
        col_idx = 1
        for header in static_headers:
            ws.cell(row=1, column=col_idx, value=header)
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
            
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = border
            col_idx += 1

        # Dynamic Payment Headers
        for i in range(1, max_payments + 1):
            # Number to ordinal (1st, 2nd, 3rd, etc.)
            suffix = ['th', 'st', 'nd', 'rd', 'th'][min(i % 10, 4)]
            if 11 <= (i % 100) <= 13:
                suffix = 'th'
            payment_title = f"{i}{suffix} PAYMENT"
            
            # Merge 3 columns for this payment
            start_col = col_idx
            end_col = col_idx + 2
            
            ws.cell(row=1, column=start_col, value=payment_title)
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            
            top_cell = ws.cell(row=1, column=start_col)
            top_cell.font = header_font
            top_cell.fill = header_fill
            top_cell.alignment = alignment_center
            top_cell.border = border

            # Row 2 sub-headers
            sub_headers = ["PAID AMOUNT", "MODE OF PAYMENT", "DATE"]
            for j, sub in enumerate(sub_headers):
                cell = ws.cell(row=2, column=start_col + j, value=sub)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = border
            
            col_idx += 3

        # Final Status Header
        ws.cell(row=1, column=col_idx, value="STATUS OF FEE")
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = border

        # Write Data
        row_idx = 3
        for item in data:
            c = 1
            ws.cell(row=row_idx, column=c, value=item['sl_no']); c+=1
            ws.cell(row=row_idx, column=c, value=item['handled_by']); c+=1
            ws.cell(row=row_idx, column=c, value=item['date_of_joining']); c+=1
            ws.cell(row=row_idx, column=c, value=item['name']); c+=1
            ws.cell(row=row_idx, column=c, value=item['ph_no']); c+=1
            ws.cell(row=row_idx, column=c, value=item['parent_name']); c+=1
            ws.cell(row=row_idx, column=c, value=item['parent_no']); c+=1
            ws.cell(row=row_idx, column=c, value=item['mail_id']); c+=1
            ws.cell(row=row_idx, column=c, value=item['qualification']); c+=1
            ws.cell(row=row_idx, column=c, value=item['campus']); c+=1
            ws.cell(row=row_idx, column=c, value=item['mode_of_study']); c+=1
            ws.cell(row=row_idx, column=c, value=item['preferred_country']); c+=1
            ws.cell(row=row_idx, column=c, value=item['preferred_level']); c+=1
            ws.cell(row=row_idx, column=c, value=item['package_chosen']); c+=1
            ws.cell(row=row_idx, column=c, value=item['total_fee']); c+=1
            ws.cell(row=row_idx, column=c, value=item['special_discount']); c+=1
            ws.cell(row=row_idx, column=c, value=item['pending']); c+=1
            ws.cell(row=row_idx, column=c, value=item['registration_fees']); c+=1

            # Payments
            for i in range(max_payments):
                if i < len(item['payments']):
                    p = item['payments'][i]
                    ws.cell(row=row_idx, column=c, value=p['amount']); c+=1
                    ws.cell(row=row_idx, column=c, value=p['method']); c+=1
                    ws.cell(row=row_idx, column=c, value=p['date']); c+=1
                else:
                    ws.cell(row=row_idx, column=c, value=""); c+=1
                    ws.cell(row=row_idx, column=c, value=""); c+=1
                    ws.cell(row=row_idx, column=c, value=""); c+=1

            ws.cell(row=row_idx, column=c, value=item['status_of_fee']); c+=1

            # Apply border to all cells in row
            for col in range(1, c):
                ws.cell(row=row_idx, column=col).border = border

            row_idx += 1

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 30) # max 30 width

        # Return File Response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Fees_Grid_Export.xlsx'
        wb.save(response)
        
        return response
