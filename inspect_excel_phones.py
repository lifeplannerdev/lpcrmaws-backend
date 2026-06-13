import openpyxl

file_path = r'b:\lp alternative\updated assets list kochi.xlsx'

print("Loading workbook...")
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"\n--- Sheet: {sheetname} ---")
    
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        # Check if "PHONE" or "IMEI" is in any cell
        if any(cell and isinstance(cell, str) and ("PHONE" in cell.upper() or "IMEI" in cell.upper() or "ASWATHY" in cell.upper()) for cell in row):
            print(f"Row {i+1}: {row}")
            # Print the next 15 rows after finding a match
            for j in range(1, 16):
                next_row = [cell.value for cell in sheet[i+1+j]]
                if any(c is not None for c in next_row):
                    print(f"Row {i+1+j}: {next_row}")
            break
